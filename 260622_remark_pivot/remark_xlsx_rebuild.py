# remark_xlsx_rebuild.py
# 2026-06-23  Jonghyun Park w/ Claude
#
# 클래식 피봇 xlsx 리마킹 전용
# 기존 xlsx 그대로 열기 → 불필요 시트 삭제 → raw 시트 셀 값만 _remark 치환 → 저장
# 피봇 캐시 XML 유지 (새 workbook 미생성) → Excel 새로고침으로 피봇 복원 가능
#
# 치환 대상: sitecode·country·subs·region + channel + ITEM(좌측열 Paid/Non-Paid 조건부)
# 치환 범위는 xlsx 안의 dim 컬럼 값 그대로 — 외부 CSV 조회 불필요

import re, random
from pathlib import Path
import openpyxl

# ════ 사용자가 바꿔야 하는 부분 ════

# ─── 출력 파일명 prefix ───
OUT_PREFIX = "_remark_"  # prefix + 원본파일명 + .xlsx  예) _remark_Analysis_260616_updated.xlsx

# ─── 경로 ───
OUT_DIR      = r"C:\Users\user_name\OneDrive - company_name\user_id\path\to\output"
CLASSIC_XLSX = r"C:\Users\user_name\Downloads\2026 CAMPAIGN NAME Campaign Performance Analysis.xlsx"

# ─── 유지할 시트 목록 (새 파일은 Excel에서 시트 탭 확인 후 교체) ───
CLASSIC_KEEP = [
    "SHEET_A TRAFFIC ANALYSIS",
    "SHEET_A_PIVOT_1",
    "SHEET_A_PIVOT_2",
    "SHEET_C PURCHASE",
    "SHEET_D PRODUCTS",
    "SHEET_A_RAW",
    "SHEET_B",
    "SHEET_C_RAW",
    "SHEET_D_RAW",
]

# ─── raw 시트별 헤더 행 번호 (데이터는 그 다음 행부터) ───
# ※ 헤더가 1행이 아닌 경우가 많음 — 새 파일은 직접 확인 필요:
#   python -c "import openpyxl; wb=openpyxl.load_workbook('파일.xlsx',read_only=True,data_only=True); ws=wb['SHEET_A_RAW']; [print(i,list(r)) for i,r in enumerate(ws.iter_rows(1,5,values_only=True),1)]"
CLASSIC_RAW_HEADER_ROW = {
    "SHEET_A_RAW": 2,   # row1 = SUBTOTAL 수식
    "SHEET_B":     8,   # row1~7 = 제목/설명
    "SHEET_C_RAW": 2,   # row1 = VLOOKUP 메모
    "SHEET_D_RAW": 1,   # row1 = 헤더
}

# ════ 내부 사용 ════

# ITEM 컬럼명
ITEM_COL_NAMES = {"item"}
# 왼쪽 열 값이 이 중 하나일 때만 ITEM 치환
ITEM_TRIGGER = {"paid", "non-paid", "non paid"}

SEED = <REMARK_SEED>

# 컬럼 헤더 → dim 여부 판단 (소문자·비알파뉴메릭 제거 후 비교)
DIM_HEADERS = {
    "sitecode", "sitecodes", "site",
    "country",
    "subs",
    "region",
    "channel",
    "channelsource",
    "channelunified",
    "variablesmarketingchannel",
    "mktchannel",
}


def _make_cipher(seed=SEED):
    rng = random.Random(seed)
    lower = list("abcdefghijklmnopqrstuvwxyz")
    shuffled = lower[:]
    while any(shuffled[i] == lower[i] for i in range(len(lower))):
        rng.shuffle(shuffled)
    return {c: s for c, s in zip(lower, shuffled)}

_LOWER_MAP = _make_cipher()
_UPPER_MAP  = {c.upper(): v.upper() for c, v in _LOWER_MAP.items()}
_CHAR_MAP   = {**_LOWER_MAP, **_UPPER_MAP}
_CACHE: dict = {}

def _mask_token(tok: str) -> str:
    if tok not in _CACHE:
        _CACHE[tok] = "".join(_CHAR_MAP.get(c, c) for c in tok)
    return _CACHE[tok]

def fx(val) -> str:
    if val is None or val == "":
        return val
    s = str(val)
    try:
        float(s); return val
    except (ValueError, TypeError):
        pass
    return "".join(
        _mask_token(p) if re.match(r"^[A-Za-z]+$", p) else p
        for p in re.split(r"([^A-Za-z]+)", s)
    )


def norm_header(h) -> str:
    return re.sub(r"[^a-z0-9]", "", str(h).lower())


def process_raw_sheet(ws, header_row_num: int = 1):
    hrow = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num))[0]
    dim_cols  = set()
    item_cols = set()

    for cell in hrow:
        h = norm_header(cell.value or "")
        if not h:
            continue
        if h in DIM_HEADERS:
            dim_cols.add(cell.column)
        if h in {norm_header(n) for n in ITEM_COL_NAMES}:
            item_cols.add(cell.column)

    if not dim_cols and not item_cols:
        print(f"    → 매핑 컬럼 없음, skip")
        return 0

    print(f"    → dim cols: {dim_cols}  item cols: {item_cols}")

    changed = 0
    total   = 0

    for row in ws.iter_rows(min_row=header_row_num + 1):
        total += 1
        row_vals = {cell.column: cell.value for cell in row}

        for cell in row:
            c = cell.column
            val = cell.value
            if val is None or not isinstance(val, str):
                continue

            if c in dim_cols:
                new_val = fx(val)
                if new_val != val:
                    cell.value = new_val
                    changed += 1
                continue

            if c in item_cols:
                left_val = str(row_vals.get(c - 1, "") or "").strip().lower()
                if left_val in ITEM_TRIGGER:
                    new_val = fx(val)
                    if new_val != val:
                        cell.value = new_val
                        changed += 1

    print(f"    → {total} rows, {changed} cells changed")
    return changed


def main():
    src = CLASSIC_XLSX
    print(f"── {Path(src).name} ──")
    wb = openpyxl.load_workbook(src)
    print(f"  Available: {wb.sheetnames}")

    for s in [s for s in wb.sheetnames if s not in CLASSIC_KEEP]:
        del wb[s]
        print(f"  Deleted: {s}")

    for sname, hrow in CLASSIC_RAW_HEADER_ROW.items():
        if sname not in wb.sheetnames:
            print(f"  SKIP (없음): {sname}")
            continue
        print(f"  Remarking: {sname} (header row={hrow})")
        process_raw_sheet(wb[sname], header_row_num=hrow)

    out_name = OUT_PREFIX + Path(src).stem + ".xlsx"
    out_path = Path(OUT_DIR) / out_name
    wb.save(str(out_path))
    print(f"\n  Saved → {out_path}")
    print(f"  Sheets: {wb.sheetnames}")

    import csv
    legend = Path(OUT_DIR) / "remark_prefix.csv"
    with open(legend, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["Token_Original", "Token_fx"])
        w.writeheader()
        for orig in sorted(_CACHE, key=str.lower):
            w.writerow({"Token_Original": orig, "Token_fx": _CACHE[orig]})
    print(f"  Legend: {legend}  ({len(_CACHE)} tokens)")
    print("\nDone.")


if __name__ == "__main__":
    main()

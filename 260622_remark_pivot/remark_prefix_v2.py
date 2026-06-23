# remark_prefix_v2.py
# 2026-06-23  Jonghyun Park w/ Claude
#
# remark_prefix.csv 의 구조화 버전 — 컬럼별 원본/fx 쌍 표시
# 샘플(remark_sample_fx.xlsx)과 동일한 컬럼쌍 레이아웃
#
# 출력: remark_prefix_v2.xlsx (2 시트)
#   Sheet "Site"    : Region | Region_fx | Subs | Subs_fx | Country | Country_fx | Site Code | Site Code_fx
#   Sheet "Channel" : channel_source | channel_source_fx | channel_unified | channel_unified_fx | paid_type

import csv, re, random
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ════ 사용자가 바꿔야 하는 부분 ════

DIM_DIR  = r"C:\Users\user_name\Downloads\data\dim"
OUT_FILE = r"C:\Users\user_name\OneDrive - company_name\user_id\path\to\output\remark_prefix_v2.xlsx"

# ════ 내부 사용 ════

SEED = <REMARK_SEED>

def _make_cipher(seed=SEED):
    rng = random.Random(seed)
    lower = list("abcdefghijklmnopqrstuvwxyz")
    shuffled = lower[:]
    while any(shuffled[i] == lower[i] for i in range(len(lower))):
        rng.shuffle(shuffled)
    return {c: s for c, s in zip(lower, shuffled)}

_LOWER_MAP = _make_cipher()
_UPPER_MAP  = {c.upper(): v.upper() for c, v in _LOWER_MAP.items()}
_CHAR_MAP   = {**_LOWER_MAP, **_UPPER_MAP}

def fx(val) -> str:
    if not val:
        return val
    s = str(val)
    try:
        float(s); return s
    except (ValueError, TypeError):
        pass
    return "".join(
        "".join(_CHAR_MAP.get(c, c) for c in p) if re.match(r"^[A-Za-z]+$", p) else p
        for p in re.split(r"([^A-Za-z]+)", s)
    )

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL    = PatternFill("solid", fgColor="DCE6F1")

def write_header(ws, headers: list):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

def write_rows(ws, rows: list, start_row=2):
    for r, row in enumerate(rows, start_row):
        fill = ALT_FILL if r % 2 == 0 else None
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if fill:
                cell.fill = fill

def auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)

def main():
    d = Path(DIM_DIR)
    wb = openpyxl.Workbook()

    # ── Sheet "Site" — d_country.csv ──
    ws_site = wb.active
    ws_site.title = "Site"

    site_headers = ["Region", "Region_fx", "Subs", "Subs_fx", "Country", "Country_fx", "Site Code", "Site Code_fx"]
    write_header(ws_site, site_headers)

    site_rows = []
    country_path = d / "d_country.csv"
    if country_path.exists():
        with open(country_path, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                region  = row.get("region", "").strip()
                subs    = row.get("subs", "").strip()
                country = row.get("country", "").strip()
                site    = row.get("sitecode", "").strip()
                site_rows.append([
                    region,  fx(region),
                    subs,    fx(subs),
                    country, fx(country),
                    site,    fx(site),
                ])
    site_rows.sort(key=lambda r: (r[0], r[2], r[4]))
    write_rows(ws_site, site_rows)
    auto_width(ws_site)
    ws_site.freeze_panes = "A2"
    print(f"Site sheet: {len(site_rows)} rows")

    # ── Sheet "Channel" — d_channel.csv ──
    ws_ch = wb.create_sheet("Channel")

    ch_headers = ["channel_source", "channel_source_fx", "channel_unified", "channel_unified_fx", "paid_type"]
    write_header(ws_ch, ch_headers)

    ch_rows = []
    channel_path = d / "d_channel.csv"
    if channel_path.exists():
        with open(channel_path, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                src     = row.get("channel_source", "").strip()
                unified = row.get("channel_unified", "").strip()
                paid    = row.get("paid_type", "").strip()
                ch_rows.append([
                    src,     fx(src),
                    unified, fx(unified),
                    paid,
                ])
    ch_rows.sort(key=lambda r: (r[4], r[0]))
    write_rows(ws_ch, ch_rows)
    auto_width(ws_ch)
    ws_ch.freeze_panes = "A2"
    print(f"Channel sheet: {len(ch_rows)} rows")

    wb.save(OUT_FILE)
    print(f"\nSaved → {OUT_FILE}")

if __name__ == "__main__":
    main()

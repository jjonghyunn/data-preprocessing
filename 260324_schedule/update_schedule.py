"""
update_schedule.py
2026-04-15  Jonghyun Park w/ Claude
2026-04-21  Jonghyun Park w/ Claude
2026-05-08  Jonghyun Park w/ Claude
2026-07-22  Jonghyun Park w/ Claude  — latest_file_key: 메일 중복 suffix _YYMMDD_HHMM(시각) 인식
2026-07-30  Jonghyun Park w/ Claude  — ① 읽기/붙여넣기 범위를 상단 상수로 추출(SRC_*/TGT_*/COMPARE)
                                       ② 대상 영역과 겹치는 병합셀 자동 해제 — MergedCell 은 value 설정이
                                          불가(read-only)해서 클리어 단계에서 예외로 죽던 문제

1. 1.고객 법인 일정 파일/ 폴더에서 최신 파일 자동 선택
   - 정렬 기준: 파일명 내 날짜(YYMMDD) → 버전(_vX.XX) → 끝 번호(_2 등) → 메일수신 일시
2. 소스 파일 첫 번째 시트 B3:J(마지막 데이터 행) 값 읽기  ※ 범위는 상단 SRC_* 상수
   - datetime → yyyy-mm-dd 문자열 변환
   - WEEKNUM 수식 셀 → W01 형식 변환
3. Auto 파일의 '고객법인일정파일' 시트 B2:K999 클리어 후 B2부터 값 붙여넣기 (서식 제외)
   ※ 범위는 상단 TGT_* 상수
"""

import re
import datetime as dt
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
import win32com.client

CHANGED_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL      = PatternFill(fill_type=None)


# ── 최신 파일 정렬 키 ────────────────────────────────────────
def mail_stamp_key(date6: str | None, hhmm: str | None) -> int:
    """메일수신 일시(YYMMDD[_HHMM])를 정렬 가능한 정수 하나로 합침.

    YYMMDD*10000 + HHMM → 260722_1432 = 2607221432, 260722(시각없음) = 2607220000
    시각 없는 옛 파일이 같은 날 시각 있는 파일보다 항상 앞(=오래된 것)으로 정렬됨.
    """
    return int(date6 or 0) * 10000 + int(hhmm or 0)


def latest_file_key(f: Path):
    """파일명에서 정렬 키 (문서날짜, HHMM, 버전float, 버전int, 메일수신일시) 반환.

    ※ 마지막 성분(메일수신일시)은 check_mail_attachment.py 가 **같은 파일명이 재수신될 때만**
      덧붙이는 suffix. _YYMMDD → _YYMMDD_HHMM (시각 포함) 으로 확장됐고,
      옛 _YYMMDD 형식도 그대로 파싱되므로 기존 파일 재정렬 문제 없음.

    A형 (8자리 날짜):
      YYYYMMDD_HHMM[_YYMMDD[_HHMM]]  → (doc_date, hhmm, 0, 0, mail)
      YYYYMMDD_vN[_YYMMDD[_HHMM]]    → (doc_date, 0, 0, ver_int, mail)
      YYYYMMDD_YYMMDD[_HHMM]         → (doc_date, 0, 0, 0, mail)
      YYYYMMDD                       → (doc_date, 0, 0, 0, 0)

    B형 (6자리 날짜 + vX.XX 버전):
      _vX.XX_YYMMDD[_YYMMDD_HHMM]    → (doc_date6, 0, ver_float, suffix, mail)
    """
    name = f.stem

    # 메일 suffix 공통 꼬리: _YYMMDD 또는 _YYMMDD_HHMM (둘 다 없어도 됨)
    MAIL_TAIL = r'(?:_(\d{6})(?:_(\d{4}))?)?'

    m8 = re.search(r'(?<!\d)(\d{8})(?!\d)', name)
    if m8:
        doc_date = int(m8.group(1))

        # YYYYMMDD_HHMM[_메일꼬리]: 뒤에 4자리 숫자가 오되 그 직후 숫자 없을 때
        m = re.search(r'(?<!\d)\d{8}_(\d{4})' + MAIL_TAIL + r'(?!\d)', name)
        if m:
            return (doc_date, int(m.group(1)), 0.0, 0, mail_stamp_key(m.group(2), m.group(3)))

        # YYYYMMDD_vN[_메일꼬리]
        m = re.search(r'(?<!\d)\d{8}_v(\d+)' + MAIL_TAIL, name)
        if m:
            return (doc_date, 0, 0.0, int(m.group(1)), mail_stamp_key(m.group(2), m.group(3)))

        # YYYYMMDD_YYMMDD[_HHMM]
        m = re.search(r'(?<!\d)\d{8}_(\d{6})(?:_(\d{4}))?(?!\d)', name)
        if m:
            return (doc_date, 0, 0.0, 0, mail_stamp_key(m.group(1), m.group(2)))

        return (doc_date, 0, 0.0, 0, 0)

    # ── B형: _vX.XX_YYMMDD ──
    # ⚠ 끝번호 정규식 `_(\d{1,5})$` 이 메일 suffix 의 시각(_1432)을 버전 끝번호로 오인하므로,
    #   끝의 `_YYMMDD_HHMM`(날짜+시각이 둘 다 있는 형태 = 스크립트가 붙인 것) 을 먼저 떼어낸 뒤 판정.
    #   날짜만 있는 `_YYMMDD` 는 문서날짜일 수 있어 떼지 않음 (종전 동작 유지).
    m_tail   = re.search(r'_(\d{6})_(\d{4})$', name)
    mail_key = mail_stamp_key(m_tail.group(1), m_tail.group(2)) if m_tail else 0
    core     = name[:m_tail.start()] if m_tail else name

    date6   = int(m.group()) if (m := re.search(r'(?<!\d)\d{6}(?!\d)', core)) else 0
    version = float(m.group(1)) if (m := re.search(r'_v(\d+\.\d+)', core)) else 0.0
    suffix  = int(m.group(1)) if (m := re.search(r'_(\d{1,5})$', core)) else 0
    return (date6, 0, version, suffix, mail_key)


# ── 경로 설정 ────────────────────────────────────────────────
BASE = Path(
    r"C:\Users\user_name\OneDrive - company_name"
    r"\Project_team_name - 1 company_name - 02 part_name"
    r"\part_name\2026\# CAMPAIGN_PROJECTS\02. CAMPAIGN NAME\02. SCHEDULE"
)

SOURCE_FOLDER    = BASE / "1.고객 법인 일정 파일"
TARGET_SHEET     = "고객법인일정파일"
LAST_SOURCE_FILE = BASE / "schedule_last_source.txt"  # 마커: Auto 파일과 같은 폴더 (프로젝트별 독립 관리. 다른 캠페인으로 fork 시 BASE만 교체하면 마커도 따라감)

# ── 읽기 / 붙여넣기 범위 ─────────────────────────────────────
# 소스 일정표의 열 구성이 바뀌면 여기 숫자만 고치면 된다 (함수 본문엔 숫자를 박지 않는다).
SRC_MIN_COL   = 2    # B (Global)
SRC_MAX_COL   = 10   # J — 1세트 구성 기준
SRC_MIN_ROW   = 3    # 소스 데이터 시작 행
TGT_START_ROW = 2    # 타겟 붙여넣기 시작 행 (소스 행 - 1)
TGT_MAX_ROW   = 999  # 클리어 범위 하단
TGT_MIN_COL   = 2    # B
TGT_MAX_COL   = 11   # K

# 전후 비교(노란 음영) 대상 — {src_data 인덱스: 타겟 열번호},  타겟 열번호 = 인덱스 + 2
COMPARE = {
    3: 5,   # E Participation
    4: 6,   # F Starts at
    6: 8,   # H Ends at
}

# ※ 소스가 B2B/B2C 2세트(B~N)로 확장된 일정표라면 위 값을 아래로 바꿔 쓴다:
#     SRC_MAX_COL = 14 (N) / TGT_MAX_COL = 14
#     COMPARE 에 { 8: 10 (J 2번째 Starts at), 10: 12 (L 2번째 Ends at) } 추가
#   (포맷이 바뀐 직후 1회는 전후 비교에서 성격이 다른 열끼리 비교돼 음영이 과하게 찍힐 수 있음)

# ── Auto 파일 자동 탐색 ──────────────────────────────────────
auto_files = list(BASE.glob("*Auto*.xlsx"))
if not auto_files:
    raise FileNotFoundError(f"Auto 파일을 찾을 수 없습니다: {BASE}")
output_file = auto_files[0]
print(f"[업데이트 대상] {output_file.name}")

# ── 소스 폴더에서 최신 파일 선택 ────────────────────────────
xlsx_files = sorted(SOURCE_FOLDER.glob("*.xlsx"), key=latest_file_key)
if not xlsx_files:
    raise FileNotFoundError(f"소스 폴더에 xlsx 파일이 없습니다: {SOURCE_FOLDER}")

source_file = xlsx_files[-1]
print(f"[소스 파일] {source_file.name}")

# 소스 파일이 이전과 동일하면 업데이트 불필요 → 스킵 (파일명 + mtime 기준)
src_mtime = int(source_file.stat().st_mtime)
current_marker = f"{source_file.name}|{src_mtime}"
if LAST_SOURCE_FILE.exists() and LAST_SOURCE_FILE.read_text(encoding="utf-8").strip() == current_marker:
    print(f"[SKIP] 소스 파일 변경 없음 ({source_file.name}), 업데이트 생략")
    exit(0)

# ── Pass 1: WEEKNUM 수식이 있는 셀 위치 파악 ─────────────────
src_wb_raw = openpyxl.load_workbook(source_file, data_only=False)
src_ws_raw = src_wb_raw.worksheets[0]

weeknum_cells = set()
for row in src_ws_raw.iter_rows(min_row=SRC_MIN_ROW, min_col=SRC_MIN_COL, max_col=SRC_MAX_COL):
    for cell in row:
        if (
            cell.value
            and isinstance(cell.value, str)
            and "WEEKNUM" in cell.value.upper()
        ):
            weeknum_cells.add((cell.row, cell.column))

src_wb_raw.close()
print(f"[WEEKNUM 셀] {len(weeknum_cells)}개 감지")

# ── Pass 2: 실제 값 읽기 ─────────────────────────────────────
src_wb = openpyxl.load_workbook(source_file, data_only=True)
src_ws = src_wb.worksheets[0]
print(f"[소스 시트] {src_ws.title}")

src_data = []
for row in src_ws.iter_rows(min_row=SRC_MIN_ROW, min_col=SRC_MIN_COL, max_col=SRC_MAX_COL):
    if all(cell.value is None for cell in row):
        continue  # 완전 빈 행 스킵
    row_data = []
    for cell in row:
        v = cell.value
        # WEEKNUM 수식 셀 → W01 형식
        if (cell.row, cell.column) in weeknum_cells and v and isinstance(v, (int, float)):
            v = f"W{int(v):02d}"
        # datetime → date로 변환 (시간 정보 제거, Excel 날짜값 유지)
        elif isinstance(v, dt.datetime):
            v = v.date()
        row_data.append(v)
    src_data.append(row_data)

src_wb.close()
print(f"[읽은 행 수] {len(src_data)}행")

# ── 이전 파일 읽기 (전후 비교용) ──────────────────────────────
prev_data = {}
if len(xlsx_files) >= 2:
    prev_file = xlsx_files[-2]
    prev_wb_raw = openpyxl.load_workbook(prev_file, data_only=False)
    prev_ws_raw = prev_wb_raw.worksheets[0]
    prev_weeknum_cells = set()
    for row in prev_ws_raw.iter_rows(min_row=SRC_MIN_ROW, min_col=SRC_MIN_COL, max_col=SRC_MAX_COL):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "WEEKNUM" in cell.value.upper():
                prev_weeknum_cells.add((cell.row, cell.column))
    prev_wb_raw.close()

    prev_wb = openpyxl.load_workbook(prev_file, data_only=True)
    prev_ws = prev_wb.worksheets[0]
    for row in prev_ws.iter_rows(min_row=SRC_MIN_ROW, min_col=SRC_MIN_COL, max_col=SRC_MAX_COL):
        if all(cell.value is None for cell in row):
            continue
        row_data = []
        for cell in row:
            v = cell.value
            if (cell.row, cell.column) in prev_weeknum_cells and v and isinstance(v, (int, float)):
                v = f"W{int(v):02d}"
            elif isinstance(v, dt.datetime):
                v = v.date()
            row_data.append(v)
        subs_key = row_data[1]  # C열
        if subs_key:
            prev_data[subs_key] = row_data
    prev_wb.close()
    print(f"[이전 파일] {prev_file.name} ({len(prev_data)}행 로드)")

# ── 타겟 파일 업데이트 ───────────────────────────────────────
try:
    tgt_wb = openpyxl.load_workbook(output_file)
except PermissionError:
    print(f"[SKIP] 파일이 사용 중입니다. 다음 실행 시 재시도합니다: {output_file.name}")
    exit(0)

if TARGET_SHEET not in tgt_wb.sheetnames:
    raise ValueError(f"'{TARGET_SHEET}' 시트를 찾을 수 없습니다. 시트 목록: {tgt_wb.sheetnames}")

tgt_ws = tgt_wb[TARGET_SHEET]

# D1에 소스 파일명 기록
tgt_ws.cell(row=1, column=4, value=source_file.name)

# 대상 영역과 겹치는 병합셀 해제 — MergedCell 은 value 설정이 불가(read-only)해서
# 클리어·붙여넣기에서 예외가 난다. 이 영역은 어차피 소스값으로 덮어쓰므로 해제해도 무방.
for rng in list(tgt_ws.merged_cells.ranges):
    if (rng.max_row >= TGT_START_ROW and rng.min_row <= TGT_MAX_ROW
            and rng.max_col >= TGT_MIN_COL and rng.min_col <= TGT_MAX_COL):
        tgt_ws.unmerge_cells(str(rng))

# 값·음영 클리어 (서식 유지)
for row in tgt_ws.iter_rows(min_row=TGT_START_ROW, max_row=TGT_MAX_ROW,
                            min_col=TGT_MIN_COL, max_col=TGT_MAX_COL):
    for cell in row:
        cell.value = None
        cell.fill  = NO_FILL

# B2부터 값 붙여넣기
for r_idx, row_data in enumerate(src_data, start=TGT_START_ROW):
    for c_idx, value in enumerate(row_data, start=TGT_MIN_COL):  # B열=2
        cell = tgt_ws.cell(row=r_idx, column=c_idx, value=value)
        if isinstance(value, dt.date):
            cell.number_format = "YYYY-MM-DD"

# 변경 셀 음영 표시 (대상 = 상단 COMPARE)
if prev_data:
    changed_count = 0
    for r_idx, row_data in enumerate(src_data, start=TGT_START_ROW):
        subs_key = row_data[1]  # C열
        if not subs_key or subs_key not in prev_data:
            continue
        prev_row = prev_data[subs_key]
        for src_idx, tgt_col in COMPARE.items():
            cur_val  = row_data[src_idx] if src_idx < len(row_data) else None
            prv_val  = prev_row[src_idx] if src_idx < len(prev_row) else None
            if cur_val != prv_val:
                tgt_ws.cell(row=r_idx, column=tgt_col).fill = CHANGED_FILL
                changed_count += 1
    print(f"[변경 셀] {changed_count}개 음영 표시")

try:
    tgt_wb.save(output_file)
    tgt_wb.close()
except PermissionError:
    tgt_wb.close()
    print(f"[SKIP] 저장 중 파일이 잠겼습니다. 다음 실행 시 재시도합니다: {output_file.name}")
    exit(0)

# Excel로 열어서 전체 재계산 후 저장 (FILTER/SORT 등 동적 배열 함수 반영)
excel = win32com.client.Dispatch("Excel.Application")
excel.Visible = False
try:
    wb_com = excel.Workbooks.Open(str(output_file.resolve()))
    excel.CalculateFull()
    wb_com.Save()
    wb_com.Close()
    LAST_SOURCE_FILE.write_text(current_marker, encoding="utf-8")
    print(f"[완료] {output_file.name} 저장 완료")
finally:
    excel.Quit()

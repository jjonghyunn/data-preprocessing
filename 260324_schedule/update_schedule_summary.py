"""
update_schedule_summary.py   [CAMPAIGN NAME 폴더 전용 — update_schedule.py 의 정제 통합판]
2026-08-11  Jonghyun Park w/ Claude
2026-08-19  Jonghyun Park w/ Claude  — public repo 공개판 (설정값 placeholder 화)

update_schedule.py 와의 차이 = **Summary 시트 자동 정제 단계가 앞에 붙었다**.

배경: 메일로 오는 고객 법인 일정 파일의 포맷이 바뀌어 이제 `Summary` 시트 하나만 온다.
      예전엔 첨부의 첫 시트가 이미 `일정`(B~N 13열) 형태라 그대로 붙여넣을 수 있었지만,
      지금은 사람이 손으로 `일정` 시트를 만들어야(기간 `6/24~9/13` 을 날짜 2개로 쪼개기 등)
      update_schedule.py 가 돌아간다. 그 수동 단계를 이 스크립트가 대신한다.

흐름:
 1. `1.고객 법인 일정 파일/` 폴더에서 최신 파일 자동 선택 (update_schedule.py 와 동일 정렬 규칙)
 2. 마커 비교 → 소스 변경 없으면 SKIP
 3. **[신규] Summary 시트 정제 → 일정 13열 데이터 생성**
 4. **[신규] 소스 xlsx 에 `일정` 시트로 기록** (WRITE_SHEET_TO_SOURCE=True 일 때)
 5. 직전 소스 파일도 같은 정제 → 전후 비교(노란 음영)용
 6. Auto 파일 `고객법인일정파일` 시트 B2:N999 클리어 후 **B5 부터** 붙여넣기
    (B5 = Region 라벨행, B6 = 헤더행, B7~ = 데이터)
 7. Excel COM 으로 전체 재계산 후 저장

정제 룰 (Summary → 일정):
  B Global      ← Summary 의 'Region'/'Global' 텍스트가 있는 열 (그룹 시작행에만 값)
  (No. 열은 건너뜀)
  C Subs        ← 헤더가 정확히 'Subs' 인 열  (5행 라벨 'Subs.' 아님)
  D Country     ← 'Country'
  E Participation ← B2B/B2C 날짜가 하나라도 파싱되면 'O', 아니면 빈칸
  F/H B2B 시작·종료 ← '캠페인 기간(B2B)' 을 '~' 로 분리 → M/D 파싱 → date
  G/I           ← 공백 (원래 WEEKNUM 자리)
  J/L B2C 시작·종료 ← '캠페인 기간(B2C)' 동일 처리
  K/M           ← 공백
  N note        ← 'Remark'

  ※ 'TBU' / '-' / 빈칸 처럼 '~' 가 없는 값은 시작·종료 모두 빈칸 처리.
  ※ 생성되는 `일정` 시트는 6행=Region 라벨, 7행=헤더, 8행~=데이터 (수동 작업본과 동일 레이아웃).
     구 update_schedule.py(SRC_MIN_ROW=3)를 돌려도 같은 결과가 나오도록 맞춘 것.
"""

import os
import re
import time
import datetime as dt
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
import win32com.client
import pywintypes


# ════════════════ 사용자가 바꿔야 하는 부분 ════════════════

# ─── 경로 ────────────────────────────────────────────────────
BASE = Path(
    r"C:\Users\user_name\OneDrive - company_name"
    r"\Project_team_name - 1 company_name - 02 part_name"
    r"\part_name\2026\# CAMPAIGN_PROJECTS\03. CAMPAIGN NAME\02. SCHEDULE"
)

SOURCE_FOLDER    = BASE / "1.고객 법인 일정 파일"
TARGET_SHEET     = "고객법인일정파일"
LAST_SOURCE_FILE = BASE / "campaign_schedule_last_source.txt"  # 마커: Auto 파일과 같은 폴더 (프로젝트별 독립 관리)

# ─── Excel COM 재계산 ────────────────────────────────────────
COM_RETRIES        = 2   # Excel 인스턴스가 도중에 죽었을 때(RPC_E_DISCONNECTED) 새 인스턴스로 재시도할 횟수
COM_RETRY_WAIT_SEC = 5   # 재시도 전 대기 (죽은 프로세스가 정리될 시간)

# ─── 정제 ────────────────────────────────────────────────────
CAMPAIGN_YEAR         = 2026        # 기간 텍스트 'M/D' 에 붙일 연도
SUMMARY_SHEET         = "Summary"   # 소스 원본 시트명 (없으면 첫 번째 시트로 fallback)
SCHEDULE_SHEET        = "일정"      # 생성할 정제 시트명
WRITE_SHEET_TO_SOURCE = True        # False 면 메모리 처리만 (소스 파일 미변경)

# Summary 헤더 문자열 — 열 문자 대신 이 이름으로 열을 찾는다 (고객이 열을 끼워넣어도 견디도록)
H_GLOBAL  = "Global"            # 이 헤더가 있는 열 = Region/Global 열
H_SUBS    = "Subs"              # 정확일치. 'Subs.'(라벨행) 는 매칭 안 됨
H_COUNTRY = "Country"
H_EPP     = "캠페인 기간(B2B)"
H_B2C     = "캠페인 기간(B2C)"
H_REMARK  = "Remark"

# 생성 시트 레이아웃 (수동 작업본과 동일)
SCHED_LABEL_ROW  = 6   # Region 라벨행
SCHED_HEADER_ROW = 7   # 헤더행 (데이터는 그 다음 행부터)

# 생성 시트 헤더행에 쓸 값 (B~N 13칸)
SCHED_HEADER = ["Global", "Subs", "Country", None,
                H_EPP, None, None, None,
                H_B2C, None, None, None,
                "note"]

# ─── 읽기/붙여넣기 범위 ───────────────────────────────────────
SRC_MIN_COL   = 2    # B (Global)
SRC_MAX_COL   = 14   # N (note)
TGT_CLEAR_ROW = 2    # 클리어 시작 행 (붙여넣기 위쪽 잔재까지 지우도록 더 위에서 시작)
TGT_START_ROW = 5    # 붙여넣기 시작 행 = Region 라벨행 → 헤더 6행, 데이터 7행~
TGT_MAX_ROW   = 999  # 클리어 범위 하단

# 전후 비교(노란 음영) 대상 — {row_data 인덱스: 타겟 열번호}
# row_data 인덱스 0=B … 12=N,  타겟 열번호 = 인덱스 + 2
COMPARE = {
    3: 5,    # E Participation
    4: 6,    # F B2B 시작
    6: 8,    # H B2B 종료
    8: 10,   # J B2C 시작
    10: 12,  # L B2C 종료
}


# ════════════════ 내부 사용 ════════════════

CHANGED_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
NO_FILL      = PatternFill(fill_type=None)

PERIOD_SEP   = "~"                                  # 전각 ～ / ∼ 도 이 문자로 정규화 후 분리
PERIOD_ALTS  = ("～", "∼", "〜")
MD_PATTERN   = re.compile(r"^\s*(\d{1,2})\s*[/.\-]\s*(\d{1,2})\s*$")   # 'M/D' (. - 구분자도 허용)
ROW_LEN      = SRC_MAX_COL - SRC_MIN_COL + 1        # 13


# ── 최신 파일 정렬 키 ────────────────────────────────────────
def mail_stamp_key(date6: str | None, hhmm: str | None) -> int:
    """메일수신 일시(YYMMDD[_HHMM])를 정렬 가능한 정수 하나로 합침.

    YYMMDD*10000 + HHMM → 260722_1432 = 2607221432, 260722(시각없음) = 2607220000
    시각 없는 옛 파일이 같은 날 시각 있는 파일보다 항상 앞(=오래된 것)으로 정렬됨.
    """
    return int(date6 or 0) * 10000 + int(hhmm or 0)


def latest_file_key(f: Path):
    """파일명에서 정렬 키 (문서날짜, HHMM, 버전float, 버전int, 메일수신일시) 반환.

    ※ 마지막 성분(메일수신일시)은 check_mail_attachment_byname.py 가 **같은 파일명이 재수신될 때만**
      덧붙이는 suffix. 2026-07-22 부터 _YYMMDD → _YYMMDD_HHMM (시각 포함) 으로 바뀌었고,
      옛 _YYMMDD 형식도 그대로 파싱되므로 기존 파일 재정렬 문제 없음.

    A형 (8자리 날짜):
      YYYYMMDD_HHMM[_YYMMDD[_HHMM]]  → (doc_date, hhmm, 0, 0, mail)
      YYYYMMDD_vN[_YYMMDD[_HHMM]]    → (doc_date, 0, 0, ver_int, mail)
      YYYYMMDD_YYMMDD[_HHMM]         → (doc_date, 0, 0, 0, mail)
      YYYYMMDD                       → (doc_date, 0, 0, 0, 0)

    B형 (6자리 날짜 + vX.XX 버전):
      _vX.XX_YYMMDD[_YYMMDD_HHMM]    → (doc_date6, 0, ver_float, suffix, mail)
    """
    name = f.stem

    # 메일 suffix 공통 꼬리: _YYMMDD 또는 _YYMMDD_HHMM (둘 다 없어도 됨)
    MAIL_TAIL = r'(?:_(\d{6})(?:_(\d{4}))?)?'

    m8 = re.search(r'(?<!\d)(\d{8})(?!\d)', name)
    if m8:
        doc_date = int(m8.group(1))

        # YYYYMMDD_HHMM[_메일꼬리]: 뒤에 4자리 숫자가 오되 그 직후 숫자 없을 때
        m = re.search(r'(?<!\d)\d{8}_(\d{4})' + MAIL_TAIL + r'(?!\d)', name)
        if m:
            return (doc_date, int(m.group(1)), 0.0, 0, mail_stamp_key(m.group(2), m.group(3)))

        # YYYYMMDD_vN[_메일꼬리]
        m = re.search(r'(?<!\d)\d{8}_v(\d+)' + MAIL_TAIL, name)
        if m:
            return (doc_date, 0, 0.0, int(m.group(1)), mail_stamp_key(m.group(2), m.group(3)))

        # YYYYMMDD_YYMMDD[_HHMM]
        m = re.search(r'(?<!\d)\d{8}_(\d{6})(?:_(\d{4}))?(?!\d)', name)
        if m:
            return (doc_date, 0, 0.0, 0, mail_stamp_key(m.group(1), m.group(2)))

        return (doc_date, 0, 0.0, 0, 0)

    # ── B형: _vX.XX_YYMMDD ──
    # ⚠ 끝번호 정규식 `_(\d{1,5})$` 이 메일 suffix 의 시각(_1432)을 버전 끝번호로 오인하므로,
    #   끝의 `_YYMMDD_HHMM`(날짜+시각이 둘 다 있는 형태 = 우리가 붙인 것) 을 먼저 떼어낸 뒤 판정.
    #   날짜만 있는 `_YYMMDD` 는 문서날짜일 수 있어 떼지 않음 (종전 동작 유지).
    m_tail   = re.search(r'_(\d{6})_(\d{4})$', name)
    mail_key = mail_stamp_key(m_tail.group(1), m_tail.group(2)) if m_tail else 0
    core     = name[:m_tail.start()] if m_tail else name

    date6   = int(m.group()) if (m := re.search(r'(?<!\d)\d{6}(?!\d)', core)) else 0
    version = float(m.group(1)) if (m := re.search(r'_v(\d+\.\d+)', core)) else 0.0
    suffix  = int(m.group(1)) if (m := re.search(r'_(\d{1,5})$', core)) else 0
    return (date6, 0, version, suffix, mail_key)


# ── Summary 정제 ─────────────────────────────────────────────
def parse_md(token: str) -> dt.date | None:
    """'8/17' 같은 M/D 토큰 → date(CAMPAIGN_YEAR, M, D). 형식이 아니면 None."""
    m = MD_PATTERN.match(token or "")
    if not m:
        return None
    try:
        return dt.date(CAMPAIGN_YEAR, int(m.group(1)), int(m.group(2)))
    except ValueError:          # 13/45 같은 오타
        return None


def parse_period(value) -> tuple[dt.date | None, dt.date | None]:
    """'6/24~9/13' → (date, date). '~' 가 없는 값(TBU, -, 빈칸)은 (None, None)."""
    if not isinstance(value, str):
        return None, None
    s = value
    for alt in PERIOD_ALTS:
        s = s.replace(alt, PERIOD_SEP)
    if PERIOD_SEP not in s:
        return None, None
    head, tail = s.split(PERIOD_SEP, 1)
    start, end = parse_md(head), parse_md(tail)
    # 12/20~1/15 처럼 해를 넘기는 기간 → 종료일만 다음 해로
    if start and end and end < start:
        end = end.replace(year=end.year + 1)
    return start, end


def find_summary_layout(ws):
    """Summary 시트에서 헤더행 위치와 필요한 열 번호를 찾아 반환.

    반환: (header_row, {논리명: 열번호})
    """
    header_row = global_col = None
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, ws.max_column + 1):
            if isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.strip() == H_GLOBAL:
                header_row, global_col = r, c
                break
        if header_row:
            break
    if not header_row:
        raise ValueError(f"'{H_GLOBAL}' 헤더를 찾을 수 없습니다 (시트: {ws.title})")

    cols = {"global": global_col}
    for key, header in (("subs", H_SUBS), ("country", H_COUNTRY),
                        ("epp", H_EPP), ("b2c", H_B2C), ("remark", H_REMARK)):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and v.strip() == header:
                cols[key] = c
                break
        else:
            raise ValueError(f"'{header}' 헤더를 찾을 수 없습니다 (시트: {ws.title}, {header_row}행)")
    return header_row, cols


def build_schedule_rows(xlsx_path: Path) -> tuple[list, list]:
    """소스 xlsx 의 Summary 를 정제해 (헤더 2행, 데이터 행들) 반환.

    각 행은 B~N 13칸 리스트. 헤더 2행 = [Region 라벨행, 컬럼 헤더행].
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[SUMMARY_SHEET] if SUMMARY_SHEET in wb.sheetnames else wb.worksheets[0]
    header_row, cols = find_summary_layout(ws)

    # Region 라벨행 = 헤더행 바로 위 (B열의 'Region')
    label_value = ws.cell(header_row - 1, cols["global"]).value if header_row > 1 else None
    label_row   = [label_value] + [None] * (ROW_LEN - 1)

    data_rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        subs    = ws.cell(r, cols["subs"]).value
        country = ws.cell(r, cols["country"]).value
        if subs in (None, "") and country in (None, ""):
            continue                                    # 데이터 행 아님

        epp_start, epp_end = parse_period(ws.cell(r, cols["epp"]).value)
        b2c_start, b2c_end = parse_period(ws.cell(r, cols["b2c"]).value)
        participation = "O" if any((epp_start, epp_end, b2c_start, b2c_end)) else None

        data_rows.append([
            ws.cell(r, cols["global"]).value,   # B Global
            subs,                               # C Subs
            country,                            # D Country
            participation,                      # E Participation
            epp_start, None, epp_end, None,     # F~I  (G/I = WEEKNUM 자리, 공백)
            b2c_start, None, b2c_end, None,     # J~M  (K/M = WEEKNUM 자리, 공백)
            ws.cell(r, cols["remark"]).value,   # N note
        ])

    wb.close()
    return [label_row, list(SCHED_HEADER)], data_rows


def write_schedule_sheet(xlsx_path: Path, label_rows: list, data_rows: list) -> None:
    """소스 xlsx 에 정제 결과를 `일정` 시트로 기록 (기존 시트가 있으면 교체).

    파일 수정 시각(mtime)은 원래대로 되돌린다 — 마커가 '메일로 받은 버전'을 가리키도록 유지하고,
    우리가 쓴 것 때문에 다음 실행이 재처리로 오인하지 않게.
    """
    orig_stat = xlsx_path.stat()
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except PermissionError:
        print(f"[알림] 소스 파일이 사용 중이라 '{SCHEDULE_SHEET}' 시트 기록을 건너뜁니다: {xlsx_path.name}")
        return

    if SCHEDULE_SHEET in wb.sheetnames:
        del wb[SCHEDULE_SHEET]
    ws = wb.create_sheet(SCHEDULE_SHEET, 0)

    for r_offset, row_data in enumerate(label_rows + data_rows):
        r_idx = SCHED_LABEL_ROW + r_offset
        for c_offset, value in enumerate(row_data):
            cell = ws.cell(row=r_idx, column=SRC_MIN_COL + c_offset, value=value)
            if isinstance(value, dt.date):
                cell.number_format = "YYYY-MM-DD"

    try:
        wb.save(xlsx_path)
    except PermissionError:
        print(f"[알림] 소스 파일 저장 실패(사용 중) — '{SCHEDULE_SHEET}' 시트 기록 생략: {xlsx_path.name}")
        return
    finally:
        wb.close()

    os.utime(xlsx_path, (orig_stat.st_atime, orig_stat.st_mtime))
    print(f"[정제 시트] '{SCHEDULE_SHEET}' 기록 완료 ({len(data_rows)}행) — {xlsx_path.name}")


# ── Auto 파일 자동 탐색 ──────────────────────────────────────
auto_files = list(BASE.glob("*Auto*.xlsx"))
if not auto_files:
    raise FileNotFoundError(f"Auto 파일을 찾을 수 없습니다: {BASE}")
output_file = auto_files[0]
print(f"[업데이트 대상] {output_file.name}")

# ── 소스 폴더에서 최신 파일 선택 ────────────────────────────
xlsx_files = sorted(SOURCE_FOLDER.glob("*.xlsx"), key=latest_file_key)
if not xlsx_files:
    raise FileNotFoundError(f"소스 폴더에 xlsx 파일이 없습니다: {SOURCE_FOLDER}")

source_file = xlsx_files[-1]
print(f"[소스 파일] {source_file.name}")

# 소스 파일이 이전과 동일하면 업데이트 불필요 → 스킵 (파일명 + mtime 기준)
src_mtime = int(source_file.stat().st_mtime)
current_marker = f"{source_file.name}|{src_mtime}"
if LAST_SOURCE_FILE.exists() and LAST_SOURCE_FILE.read_text(encoding="utf-8").strip() == current_marker:
    print(f"[SKIP] 소스 파일 변경 없음 ({source_file.name}), 업데이트 생략")
    exit(0)

# ── Summary 정제 ─────────────────────────────────────────────
label_rows, data_rows = build_schedule_rows(source_file)
src_data = label_rows + data_rows
print(f"[정제 완료] 데이터 {len(data_rows)}행 (+ 헤더 {len(label_rows)}행)")

if WRITE_SHEET_TO_SOURCE:
    write_schedule_sheet(source_file, label_rows, data_rows)

# ── 이전 파일 정제 (전후 비교용) ─────────────────────────────
# 키가 Subs 만이면 SENA(북유럽 4국) 처럼 같은 Subs 가 여러 행인 경우 마지막 행만 남는다.
# → (Subs, Country, 같은 조합의 몇 번째) 로 키를 잡아 행 단위로 정확히 대응시킨다.
def compare_key(row_data, seen: dict):
    key = (row_data[1], row_data[2])
    seen[key] = seen.get(key, 0) + 1
    return key + (seen[key],)


prev_data = {}
if len(xlsx_files) >= 2:
    prev_file = xlsx_files[-2]
    try:
        _, prev_rows = build_schedule_rows(prev_file)
    except (ValueError, KeyError) as e:      # 옛 포맷(Summary 없음/헤더 다름) → 비교 생략
        prev_rows = []
        print(f"[알림] 이전 파일 정제 불가 — 전후 비교 생략 ({prev_file.name}): {e}")
    seen_prev = {}
    for row_data in prev_rows:
        prev_data[compare_key(row_data, seen_prev)] = row_data
    if prev_data:
        print(f"[이전 파일] {prev_file.name} ({len(prev_data)}행 로드)")

# ── 타겟 파일 업데이트 ───────────────────────────────────────
try:
    tgt_wb = openpyxl.load_workbook(output_file)
except PermissionError:
    print(f"[SKIP] 파일이 사용 중입니다. 다음 실행 시 재시도합니다: {output_file.name}")
    exit(0)

if TARGET_SHEET not in tgt_wb.sheetnames:
    raise ValueError(f"'{TARGET_SHEET}' 시트를 찾을 수 없습니다. 시트 목록: {tgt_wb.sheetnames}")

tgt_ws = tgt_wb[TARGET_SHEET]

# D1에 소스 파일명 기록
tgt_ws.cell(row=1, column=4, value=source_file.name)

# 대상 영역(B2:N999)과 겹치는 병합셀 해제 — MergedCell 은 value 설정 불가(read-only)라
# 클리어/붙여넣기에서 충돌. 어차피 이 영역은 소스값으로 덮어쓰므로 해제해도 무방.
for rng in list(tgt_ws.merged_cells.ranges):
    if (rng.max_row >= TGT_CLEAR_ROW and rng.min_row <= TGT_MAX_ROW
            and rng.max_col >= SRC_MIN_COL and rng.min_col <= SRC_MAX_COL):
        tgt_ws.unmerge_cells(str(rng))

# B2:N999 값·음영 클리어 (서식 유지)
for row in tgt_ws.iter_rows(min_row=TGT_CLEAR_ROW, max_row=TGT_MAX_ROW,
                            min_col=SRC_MIN_COL, max_col=SRC_MAX_COL):
    for cell in row:
        cell.value = None
        cell.fill  = NO_FILL

# B5(라벨행)부터 값 붙여넣기 → 헤더 6행, 데이터 7행~
for r_idx, row_data in enumerate(src_data, start=TGT_START_ROW):
    for c_idx, value in enumerate(row_data, start=SRC_MIN_COL):  # B열=2
        cell = tgt_ws.cell(row=r_idx, column=c_idx, value=value)
        if isinstance(value, dt.date):
            cell.number_format = "YYYY-MM-DD"

# 변경 셀 음영 표시 (대상 = 상단 COMPARE). 헤더 2행은 건너뛰고 데이터 행만 비교.
if prev_data:
    changed_count = 0
    seen_cur = {}
    data_start_row = TGT_START_ROW + len(label_rows)
    for r_idx, row_data in enumerate(data_rows, start=data_start_row):
        key = compare_key(row_data, seen_cur)
        if key not in prev_data:
            continue
        prev_row = prev_data[key]
        for src_idx, tgt_col in COMPARE.items():
            cur_val = row_data[src_idx] if src_idx < len(row_data) else None
            prv_val = prev_row[src_idx] if src_idx < len(prev_row) else None
            if cur_val != prv_val:
                tgt_ws.cell(row=r_idx, column=tgt_col).fill = CHANGED_FILL
                changed_count += 1
    print(f"[변경 셀] {changed_count}개 음영 표시")

try:
    tgt_wb.save(output_file)
    tgt_wb.close()
except PermissionError:
    tgt_wb.close()
    print(f"[SKIP] 저장 중 파일이 잠겼습니다. 다음 실행 시 재시도합니다: {output_file.name}")
    exit(0)

# Excel로 열어서 전체 재계산 후 저장 (FILTER/SORT 등 동적 배열 함수 반영)
# ※ openpyxl 저장은 수식의 캐시값을 지우므로 이 재계산 저장이 반드시 성공해야 한다.
#   (캐시값이 없으면 openpyxl data_only=True 로 읽는 후속 도구가 전부 None 을 본다)
# ※ DispatchEx = 전용 인스턴스. Dispatch 는 이미 떠 있는 Excel 에 붙어서, 그쪽이 먼저
#   종료되면 Quit 단계에서 AttributeError 로 죽는다(작업 스케줄러가 실패로 기록).
def recalc_and_save(path: Path) -> None:
    """Excel COM 으로 전체 재계산 후 저장. 실패 시 예외를 올린다."""
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False   # Close/Quit 시 저장 확인 팝업 방지 (스케줄러 실행 대비)
    try:
        wb_com = excel.Workbooks.Open(str(path.resolve()))
        excel.CalculateFull()
        wb_com.Save()
        # ※ gen_py 캐시가 없으면(파이썬 새 버전 설치 직후 등) win32com 이 지연 바인딩으로 동작해
        #   `wb_com.Close` 는 속성 접근만으로 COM 메서드가 실행되고 bool(True) 을 돌려준다
        #   → 이어지는 `()` 가 TypeError: 'bool' object is not callable 로 죽는다.
        #   그 시점엔 이미 저장·닫기가 끝난 상태이므로 TypeError 만 무시한다.
        #   (조기 바인딩이면 아래 호출이 정상 동작 — 양쪽 다 안전)
        try:
            wb_com.Close(SaveChanges=False)
        except TypeError:
            pass
    finally:
        try:
            excel.Quit()
        except Exception as e:      # 이미 죽은 인스턴스 등 — 저장 성패는 위에서 판정
            print(f"[알림] Excel 종료 중 무시된 예외: {e}")


# ※ 재시도 이유: Excel 인스턴스가 재계산 도중 죽으면 이후 호출이
#   com_error(-2147417848, RPC_E_DISCONNECTED '호출된 개체가 연결이 끊겼습니다') 로 실패한다.
#   프로세스 단위 사고라 같은 인스턴스로는 복구가 안 되고, 새 인스턴스로 다시 열면 대개 성공한다.
#   여기서 끝까지 실패하면 Auto 파일은 **수식 캐시가 빈 상태**로 남으므로 그 사실을 크게 알린다.
for attempt in range(1, COM_RETRIES + 1):
    try:
        recalc_and_save(output_file)
        LAST_SOURCE_FILE.write_text(current_marker, encoding="utf-8")
        print(f"[완료] {output_file.name} 저장 완료")
        break
    except pywintypes.com_error as e:
        print(f"[재시도 {attempt}/{COM_RETRIES}] Excel COM 실패: {e}")
        if attempt == COM_RETRIES:
            print(f"[실패] Excel 재계산·저장이 {COM_RETRIES}회 모두 실패했습니다.")
            print(f"        {output_file.name} 은 지금 **수식 캐시가 빈 상태**입니다 — "
                  f"Excel 로 한 번 열었다가 저장하거나 이 스크립트를 다시 실행하세요.")
            print(f"        (마커를 기록하지 않았으므로 다음 실행이 같은 소스를 재처리합니다)")
            raise
        time.sleep(COM_RETRY_WAIT_SEC)

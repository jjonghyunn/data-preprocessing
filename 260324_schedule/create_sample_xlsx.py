# -*- coding: utf-8 -*-
"""
create_sample_xlsx.py
샘플 md_schedule.xlsx 생성 스크립트

익명화 처리:
  - 법인명: CORP-XX 형식 (예: CORP-UK, CORP-WA, CORP-BN, CORP-SA)
  - 담당자: 홍길동, 김철수 등 대체명
  - 나라/site_code: 실제값 유지
  - 한총: 내부 용어 그대로 유지
"""
import openpyxl
import os

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "md_schedule"

# ── 0. 제목 ──────────────────────────────────────────────────────────────────
ws["A1"] = "sample_md_schedule (법인명 익명화)"
ws["A2"] = "* 법인명은 CORP-XX 형식으로, 담당자는 가명으로 대체됨"

# ── 1. 헤더 (Row 3) ───────────────────────────────────────────────────────────
headers = {
    # ① 최종 출력 A:H
    1: "#",  2: "RHQ",  3: "Subs",  4: "Country",  5: "RS",  6: "site_code",
    7: "start_date",  8: "end_date",
    # ② 중간 연산 M:T + U
    13: "Tier",  14: "RHQ",  15: "Subs",  16: "Country",  17: "RS",
    18: "site_code",  19: "start_date",  20: "end_date",  21: "추천제외",
    # ③ Site 마스터 Z:AH
    26: "#",  27: "RHQ",  28: "Subsidiary",  29: "Country",  30: "SiteCode",
    31: "URL",  32: "Platform",  33: "App",  34: "RS",
    # ④ 헬퍼 AL:AP
    38: "AR_clean",  39: "AR_key",  40: "Participation",
    41: "start_raw",  42: "end_raw",
    # AS 체크
    45: "사용 안됨 체크",
    # ⑤ 고객 파일 붙여넣기 AT:BB
    46: "Region",  47: "Subs",  48: "PIC",  49: "Participation",
    50: "Starts at",  51: "(Week)",  52: "Ends at",  53: "(Week)",  54: "Note",
}
for col, hdr in headers.items():
    ws.cell(3, col, hdr)

# ── 2. Site 마스터 데이터 (Z:AG), rows 4-11 ───────────────────────────────────
# (row, #, RHQ, Subsidiary(AB), Country, SiteCode(AD), URL, Platform, App)
master_rows = [
    (4,   1, "EU", "CORP-UK",  "United Kingdom",  "uk",        "example.com/uk/",        "HQ Shop", "O"),
    (5,   2, "EU", "CORP-UK",  "Ireland",         "ie",        "example.com/ie/",        "HQ Shop", "O"),
    (6,   3, "EU", "CORP-BN",  "Belgium",         "be",        "example.com/be/",        "HQ Shop", "X"),
    (7,   4, "AF", "CORP-WA",  "Africa",          "africa_en", "example.com/africa-en/", "HQ Shop", "X"),
    (8,   5, "AF", "CORP-WA",  "Africa",          "africa_fr", "example.com/africa-fr/", "HQ Shop", "X"),
    (9,   6, "AF", "CORP-WA",  "Africa",          "africa_pt", "example.com/africa-pt/", "HQ Shop", "X"),
    (10,  7, "AP", "한총",     "Korea",           "kr",        "example.com/kr/",        "HQ Shop", "O"),
    (11,  8, "AF", "CORP-SA",  "South Africa",    "za",        "example.com/za/",        "HQ Shop", "X"),
]
for row, num, rhq, subs, country, site, url, platform, app in master_rows:
    ws.cell(row, 26, num)
    ws.cell(row, 27, rhq)
    ws.cell(row, 28, subs)
    ws.cell(row, 29, country)
    ws.cell(row, 30, site)
    ws.cell(row, 31, url)
    ws.cell(row, 32, platform)
    ws.cell(row, 33, app)

# ── 3. 고객 파일 데이터 (AT:BB) ───────────────────────────────────────────────
# (row, Region, AU_Subs, PIC, Participation, Start, StartWk, End, EndWk, Note)
customer_rows = [
    (4,  "EU Region",  "Total",                  "",       "",  "",            "",    "",            "",    ""),
    (5,  "",           "CORP-UK",                "홍길동", "O", "2026-04-01",  "W14", "2026-05-31",  "W22", ""),
    (6,  "",           "CORP-UK-IE",             "김철수", "O", "2026-04-15",  "W16", "2026-05-31",  "W22", ""),
    (7,  "",           "CORP-BN-BE",             "이영희", "X", "",            "",    "",            "",    "Not participating"),
    (8,  "AF Region",  "Total",                  "",       "",  "",            "",    "",            "",    ""),
    (9,  "",           "CORP-WA (Africa_EN)",    "박민준", "O", "2026-04-01",  "W14", "2026-05-31",  "W22", ""),
    (10, "",           "CORP-WA (Africa_FR)",    "최수진", "X", "",            "",    "",            "",    ""),
    (11, "AP Region",  "Total",                  "",       "",  "",            "",    "",            "",    ""),
    (12, "",           "한총",                   "정다영", "O", "2026-04-08",  "W15", "2026-05-31",  "W22", ""),
    (13, "",           "CORP-SA (ZA)",           "강민서", "",  "",            "",    "",            "",    "TBC"),
]
for row, region, subs_au, pic, part, start, startw, end, endw, note in customer_rows:
    if region:   ws.cell(row, 46, region)
    if subs_au:  ws.cell(row, 47, subs_au)
    if pic:      ws.cell(row, 48, pic)
    if part:     ws.cell(row, 49, part)
    if start:    ws.cell(row, 50, start)
    if startw:   ws.cell(row, 51, startw)
    if end:      ws.cell(row, 52, end)
    if endw:     ws.cell(row, 53, endw)
    if note:     ws.cell(row, 54, note)

# ── 4. 헬퍼 수식 ─────────────────────────────────────────────────────────────

# AL: AR_clean (rows 4-13 — AU 데이터가 있는 전체 범위)
for r in range(4, 14):
    ws[f"AL{r}"] = (
        f'=IF(OR(NOT(ISBLANK(AT{r})),ISBLANK(AU{r}),AU{r}="Total"),"",IF(ISNUMBER(FIND("-",AU{r})),'
        f'LEFT(AU{r},FIND("-",AU{r})-1),IF(ISNUMBER(FIND(" ",AU{r})),'
        f'LEFT(AU{r},FIND(" ",AU{r})-1),AU{r})))'
    )

# 마스터 rows에 AM / AN / AO / AP / AH 수식 + M:T 블록 수식 작성
for row, num, rhq, subs, country, site, url, platform, app in master_rows:
    r = row

    # AM: AR_key (7단계 우선순위)
    ws[f"AM{r}"] = (
        f'=IF(ISBLANK(AD{r}),"",IF(COUNTIF($AU$4:$AU$200,AB{r}&"-"&UPPER(IF(ISNUMBER(FIND("_",AD{r})),LEFT(AD{r},FIND("_",AD{r})-1),AD{r})))>0,'
        f'AB{r}&"-"&UPPER(IF(ISNUMBER(FIND("_",AD{r})),LEFT(AD{r},FIND("_",AD{r})-1),AD{r})),'
        f'IF(COUNTIF($AU$4:$AU$200,AB{r}&" ("&PROPER(LEFT(AD{r},FIND("_",AD{r})-1))&"_"&UPPER(MID(AD{r},FIND("_",AD{r})+1,20))&")")>0,'
        f'AB{r}&" ("&PROPER(LEFT(AD{r},FIND("_",AD{r})-1))&"_"&UPPER(MID(AD{r},FIND("_",AD{r})+1,20))&")",'
        f'IF(COUNTIF($AU$4:$AU$200,AB{r})>0,AB{r},'
        f'IF(ISNUMBER(FIND("_",AD{r})),"",'
        f'IFERROR(INDEX($AU$4:$AU$200,MATCH(AB{r},$AL$4:$AL$200,0)),'
        f'IF(ISNUMBER(FIND("-",AB{r})),IFERROR(INDEX($AU$4:$AU$200,MATCH(LEFT(AB{r},FIND("-",AB{r})-1),$AU$4:$AU$200,0)),""),'
        f'IF(AB{r}="CORP-CA",IF(RIGHT(AD{r},3)="_fr",IFERROR(INDEX($AU$4:$AU$200,MATCH("CORP-WA (Africa_FR)",$AU$4:$AU$200,0)),""),'
        f'IF(RIGHT(AD{r},3)="_en",IFERROR(INDEX($AU$4:$AU$200,MATCH("CORP-WA (Africa_EN)",$AU$4:$AU$200,0)),""),"")),"")))))))))'
    )

    # AN: Participation
    ws[f"AN{r}"] = (
        f'=IF(LEN(TRIM(AM{r}))=0,"",IFERROR(INDEX($AW$4:$AW$200,MATCH(AM{r},$AU$4:$AU$200,0)),\"\"))'
        .replace('\\"', '"')  # 작은따옴표 문자열 내 큰따옴표 보정 불필요하지만 안전하게
    )
    # AN: Participation (clean version)
    ws[f"AN{r}"] = f'=IF(LEN(TRIM(AM{r}))=0,"",IFERROR(INDEX($AW$4:$AW$200,MATCH(AM{r},$AU$4:$AU$200,0)),""))'

    # AO: start_raw
    ws[f"AO{r}"] = f'=IF(OR(LEN(TRIM(AM{r}))=0,AN{r}="X",AN{r}<>"O"),"",IFERROR(INDEX($AX$4:$AX$200,MATCH(AM{r},$AU$4:$AU$200,0)),""))'

    # AP: end_raw
    ws[f"AP{r}"] = f'=IF(OR(LEN(TRIM(AM{r}))=0,AN{r}="X",AN{r}<>"O"),"",IFERROR(INDEX($AZ$4:$AZ$200,MATCH(AM{r},$AU$4:$AU$200,0)),""))'

    # AH: RS label
    ws[f"AH{r}"] = f'=IF(ISBLANK(AD{r}),"",IF(ISNUMBER(FIND("_",AD{r})),AC{r}&" ("&MID(AD{r},FIND("_",AD{r})+1,20)&")",AC{r}))'

    # ② M:T 중간 연산 블록
    ws[f"N{r}"] = f'=IF(ISBLANK(AD{r}),"",AA{r})'
    ws[f"O{r}"] = f'=IF(ISBLANK(AD{r}),"",IF(AM{r}="한총","CORP-KR",IF(LEN(TRIM(AM{r}))=0,AB{r},AM{r})))'
    ws[f"P{r}"] = f'=IF(ISBLANK(AD{r}),"",AC{r})'
    ws[f"Q{r}"] = f'=IF(ISBLANK(AD{r}),"",AH{r})'
    ws[f"R{r}"] = f'=IF(ISBLANK(Z{r}),"",AD{r})'
    ws[f"S{r}"] = f'=IF(ISBLANK(AD{r}),"",IF(AN{r}="X","X",AO{r}))'
    ws[f"T{r}"] = f'=IF(ISBLANK(AD{r}),"",IF(AN{r}="X","",AP{r}))'

# ── 5. AS: 사용 안됨 체크 ─────────────────────────────────────────────────────
# 고객 데이터 행 중 Region/Total 제외
check_rows = [5, 6, 7, 9, 10, 12, 13]
for r in check_rows:
    ws[f"AS{r}"] = f'=IF((AW{r}="X")+(AW{r}=""),"",IF(COUNTIF($O$4:$O$191,IF(AU{r}="한총","CORP-KR",AU{r}))>0,"","사용 안됨"))'

# ── 6. ① 최종 출력 수식 A4 ───────────────────────────────────────────────────
ws["A4"] = '=SORT(UNIQUE(FILTER(M4:T200,(Q4:Q200<>"")*(S4:S200<>"")*(U4:U200<>"O"))))'

# ── 7. 저장 ─────────────────────────────────────────────────────────────────
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "sample_md_schedule.xlsx")
wb.save(out_path)
print("Saved:", out_path)

# ── 8. 검증 ─────────────────────────────────────────────────────────────────
wb2 = openpyxl.load_workbook(out_path)
ws2 = wb2.active
print("\n=== 마스터 확인 (Z:AD) ===")
for r in range(4, 12):
    z  = ws2.cell(r, 26).value
    ab = ws2.cell(r, 28).value
    ad = ws2.cell(r, 30).value
    print(f"  row{r}: #{z} AB={ab} site={ad}")

print("\n=== 고객 데이터 (AU) ===")
for r in range(4, 14):
    at = ws2.cell(r, 46).value
    au = ws2.cell(r, 47).value
    aw = ws2.cell(r, 49).value
    print(f"  row{r}: Region={at!r} Subs={au!r} Part={aw!r}")

print("\n=== 헬퍼 수식 확인 (AM, AN, AO, AP) ===")
for r in range(4, 12):
    am = ws2.cell(r, 39).value or ''
    print(f"  row{r} AM(끝50자): ...{am[-50:]}")

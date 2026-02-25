import os
import re
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

# 📁 경로 설정
base_dir = r'C:\Users\CNXK\Downloads'

# 📄 파일명 지정
csv_filename = 'hd_qry_260108v2_us_convert.csv'
excel_filename = 'hd_campaign_main_default_value_260108_203922.xlsx'

# 📌 경로 조합
csv_path = os.path.join(base_dir, csv_filename)
excel_path = os.path.join(base_dir, excel_filename)

# ✅ 파일 존재 확인
if not os.path.exists(csv_path):
    raise FileNotFoundError(f"❌ CSV 파일이 없습니다: {csv_path}")
if not os.path.exists(excel_path):
    raise FileNotFoundError(f"❌ 엑셀 파일이 없습니다: {excel_path}")

# ✅ 날짜_시간 패턴 추출
def extract_suffix(filename):
    match = re.search(r'(\d{6}_\d{6})', filename)
    if not match:
        raise ValueError("❌ 파일명에서 'YYMMDD_HHMMSS' 패턴을 찾을 수 없습니다.")
    return match.group(1)

# ⏱️ 패턴 추출
date_time_suffix = extract_suffix(excel_filename)
sheet_name = f'output_{date_time_suffix}'
output_filename = f'campaign_main_value_mapping_{date_time_suffix}.xlsx'

# 📥 CSV 로드
# A: key(0), E: value(4), F: value_origin(5), H: start_date(7), I: end_date(8)
csv_df = pd.read_csv(csv_path, encoding='utf-8')

# 문자열 정리 (필요한 컬럼만)
csv_df.iloc[:, 0] = csv_df.iloc[:, 0].astype(str).str.strip()  # key
csv_df.iloc[:, 4] = csv_df.iloc[:, 4].astype(str).str.strip()  # value
csv_df.iloc[:, 5] = csv_df.iloc[:, 5].astype(str).str.strip()  # value_origin
csv_df.iloc[:, 7] = csv_df.iloc[:, 7].astype(str).str.strip()  # start_date
csv_df.iloc[:, 8] = csv_df.iloc[:, 8].astype(str).str.strip()  # end_date

# 🔁 키-값 dict 구성 (중복키는 list로 쌓아서 여러값일치 처리)
csv_map_E = defaultdict(list)  # value
csv_map_F = defaultdict(list)  # value_origin
csv_map_H = defaultdict(list)  # start_date
csv_map_I = defaultdict(list)  # end_date

for _, r in csv_df.iterrows():
    key = r.iloc[0]
    csv_map_E[key].append(r.iloc[4])
    csv_map_F[key].append(r.iloc[5])
    csv_map_H[key].append(r.iloc[7])
    csv_map_I[key].append(r.iloc[8])

# 📊 엑셀 로드 및 시트 열기
wb = load_workbook(excel_path)
if sheet_name not in wb.sheetnames:
    raise ValueError(f"❌ 시트 '{sheet_name}'을 찾을 수 없습니다.")
ws = wb[sheet_name]

# 열 번호 (1-based)
col_N = 14  # 참조 키 (N열)
col_M = 13  # CSV E(value)      → m열
col_O = 15  # CSV F(valueorigin)→ o열
col_P = 16  # CSV H(start_date) → p열
col_Q = 17  # CSV I(end_date)   → q열

def auto_cast_value(val):
    """숫자처럼 생긴 문자열은 int/float로 변환, 아니면 원값 유지"""
    try:
        if val == "" or val is None:
            return ""
        if isinstance(val, str):
            val = val.strip()
        if val == "":
            return ""
        f = float(val)
        if f.is_integer():
            return int(f)
        return f
    except:
        return val

def map_value(row_idx, target_col, data_map, key):
    """key 기준으로 data_map에서 값을 찾아 target_col에 기록"""
    cell = ws.cell(row=row_idx, column=target_col)
    if not key or key not in data_map:
        cell.value = ""
    elif len(data_map[key]) > 1:
        cell.value = "여러값일치"
    else:
        cell.value = auto_cast_value(data_map[key][0])

# 🔄 행 순회하며 매핑
matched_value = 0  # value(E→K) 매칭 카운트(원하면 다른 컬럼도 따로 셀 수 있음)

for row in range(2, ws.max_row + 1):
    key_cell = ws.cell(row=row, column=col_N)
    key = str(key_cell.value).strip() if key_cell.value is not None else None

    # value(E) → K
    before = ws.cell(row=row, column=col_M).value
    map_value(row, col_M, csv_map_E, key)
    # 매칭 카운트(중복값/빈값 제외하고 단일값이 들어간 경우만)
    if key and key in csv_map_E and len(csv_map_E[key]) == 1:
        matched_value += 1

    # value_origin(F) → o
    map_value(row, col_O, csv_map_F, key)

    # start_date(H) → p
    map_value(row, col_P, csv_map_H, key)

    # end_date(I) → q
    map_value(row, col_Q, csv_map_I, key)

# 💾 저장
output_path = os.path.join(base_dir, output_filename)
wb.save(output_path)
print(f"✅ 저장 완료: {output_path} (value 매칭: {matched_value}개)")

